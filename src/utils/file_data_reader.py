import xlrd
import openpyxl
import csv
from typing import Iterable, List, Any

class FileDataReader:
    def __init__(self, filePath: str):
        self.filePath = filePath

    def count(self) -> int:
        """Conta o número de linhas de dados no arquivo, excluindo o cabeçalho."""

        if self.filePath.endswith('.xls'):
            return self._countXlsLines()
        elif self.filePath.endswith('.xlsx'):
            return self._countXlsxLines()
        elif self.filePath.endswith('.csv'):
            return self._countCsvLines()
        else:
            raise ValueError("Formato de arquivo não suportado. Use [.csv, .xls, .xlsx] com cabeçalho")

    def stream(self, batchSize: int) -> Iterable[List[List[str]]]:
        """Gera lotes (batches) de linhas do arquivo."""
        if not self.filePath:
            raise ValueError("É necessário o caminho do arquivo para ler")
        if not batchSize or batchSize <= 0:
            raise ValueError("O tamanho do lote (batch) tem que ser maior que 0")

        if self.filePath.endswith('.xls'):
            yield from self._streamXls(batchSize)
        elif self.filePath.endswith('.xlsx'):
            yield from self._streamXlsx(batchSize)
        elif self.filePath.endswith('.csv'):
            yield from self._streamCsv(batchSize)
        else:
            raise ValueError("Formato de arquivo não suportado. Use [.csv, .xls, .xlsx] com cabeçalho")

    def _iteratorXls(self, rows) -> Iterable[List[Any]]:
        for row_index in range(rows.nrows):
            yield rows.row_values(row_index)
        
    def _streamXls(self, batchSize: int) -> Iterable[List[List[str]]]:
        """Lê um arquivo .xls em lotes."""
        workbook = xlrd.open_workbook(self.filePath)
        sheet = workbook.sheet_by_index(0)
        
        yield from self._generateBatches(self._iteratorXls(sheet), batchSize, skipHeader=True)

    def _streamXlsx(self, batchSize: int) -> Iterable[List[List[str]]]:
        """Lê um arquivo .xlsx em lotes."""
        workbook = openpyxl.load_workbook(self.filePath)
        sheet = workbook.active
        yield from self._generateBatches(sheet.iter_rows(values_only=True), batchSize, skipHeader=True)

    def _streamCsv(self, batchSize: int) -> Iterable[List[List[str]]]:
        """Lê um arquivo .csv em lotes."""
        with open(self.filePath, mode='r', newline='') as csvfile:
            reader = csv.reader(csvfile, delimiter=';', quotechar='|')
            yield from self._generateBatches(reader, batchSize, skipHeader=True)

    def _generateBatches(self, rows: Iterable, batchSize: int, skipHeader: bool = False) -> Iterable[List[List[str]]]:
        """Gera lotes de linhas a partir de um iterável de linhas."""
        batch = []
        for i, row in enumerate(rows):
            if skipHeader and i == 0:
                continue  # Pula o cabeçalho
            batch.append(list(row))
            if len(batch) == batchSize:
                yield batch
                batch = []
        if batch:
            yield batch

    def _countXlsLines(self) -> int:
        """Conta o número de linhas de dados em um arquivo .xls."""
        workbook = xlrd.open_workbook(self.filePath)
        sheet = workbook.sheet_by_index(0)
        return max(0, sheet.nrows - 1)  # Exclui o cabeçalho

    def _countXlsxLines(self) -> int:
        """Conta o número de linhas de dados em um arquivo .xlsx."""
        workbook = openpyxl.load_workbook(self.filePath)
        sheet = workbook.active
        return max(0, sheet.max_row - 1)  # Exclui o cabeçalho

    def _countCsvLines(self) -> int:
        """Conta o número de linhas de dados em um arquivo .csv."""
        with open(self.filePath, mode='r', newline='') as csvfile:
            reader = csv.reader(csvfile, delimiter=';', quotechar='|')
            next(reader)  # Pula o cabeçalho
            return sum(1 for _ in reader)